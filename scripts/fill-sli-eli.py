# -*- coding: utf-8 -*-
"""
Bridge: fill SLI/ELI template cells using openpyxl (xlsm -> xlsx, Office-compatible).
Reads a JSON payload from stdin:
{
  "template": "path/to/template.xlsm",
  "out": "path/to/output.xlsx",
  "sheet": "air",
  "cells": { "D23": "157-87751101", "D72": "2026-08-03", "M16": "consignee" }
}
"""
import sys
import json
from datetime import datetime, date
from openpyxl import load_workbook


def parse_value(v):
    """Convert JSON value to openpyxl-friendly value."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        # Try ISO date strings used by the workflow (YYYY-MM-DD)
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        return v
    return v


def main():
    payload = json.load(sys.stdin)
    template = payload["template"]
    out_path = payload["out"]
    sheet_name = payload["sheet"]
    cells = payload.get("cells", {})

    wb = load_workbook(template, data_only=False)  # keep formulas & styles
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet '{sheet_name}' not found in template", file=sys.stderr)
        sys.exit(1)
    ws = wb[sheet_name]

    for coord, raw in cells.items():
        ws[coord] = parse_value(raw)

    # openpyxl cannot keep macros (vbaProject.bin) -> save as xlsx (macro removed)
    wb.save(out_path)
    print(f"Written: {out_path}")


if __name__ == "__main__":
    main()