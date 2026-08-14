# -*- coding: utf-8 -*-
"""Analyze source XLS/XLSX files structure for the AGL MAWB workflow.

Usage:
    python scripts/analyze-xls.py "<path-to-xlsx>" [limit_rows]
"""
import sys
from openpyxl import load_workbook


def dump_sheet(ws, max_cols=30, limit=None):
    print(f"\n=== Sheet: {ws.title} | dims={ws.dimensions} | max_row={ws.max_row} max_col={ws.max_column} ===")
    print(f"--- Merged cells: {[str(r) for r in ws.merged_cells.ranges]} ---")

    widths = {}
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            widths[col_letter] = round(dim.width, 1)
    if widths:
        print(f"--- Column widths: {widths} ---")

    max_row = ws.max_row if limit is None else min(ws.max_row, limit)
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            max_col=min(ws.max_column, max_cols)):
        values = []
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str):
                v = repr(v)
            elif isinstance(v, float):
                v = f"{v:.4g}"
            values.append(f"{cell.coordinate}={v}")
        if values:
            print(" | ".join(values))


def main(path, limit=None, max_cols=30):
    wb = load_workbook(path, data_only=True)
    print(f"Workbook: {path}")
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        dump_sheet(ws, limit=limit, max_cols=max_cols)


if __name__ == "__main__":
    path = sys.argv[1]
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
    max_cols = int(sys.argv[3]) if len(sys.argv) > 3 else 30
    main(path, limit, max_cols)
